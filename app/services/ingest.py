import csv
import io
import logging
from typing import List

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from pypdf import PdfReader
from starlette import status

from app.models.schemas import KnowledgeItem

logger = logging.getLogger(__name__)


class IngestService:
    """
    Parse uploaded files into KB chunks to be embedded and stored.
    Supported: pdf, txt, md, csv, tsv, xlsx/xls.
    """

    def __init__(self, chunk_size: int = 800, chunk_overlap: int = 100) -> None:
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap
        self.allowed_types = {
            "application/pdf",
            "text/plain",
            "text/markdown",
            "text/csv",
            "text/tab-separated-values",
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }

    def _chunk_text(self, text: str) -> List[str]:
        words = text.split()
        if not words:
            return []
        chunks = []
        start = 0
        while start < len(words):
            end = start + self.chunk_size
            chunk = " ".join(words[start:end])
            chunks.append(chunk)
            start = end - self.chunk_overlap
            if start < 0:
                start = 0
        return chunks

    def _parse_pdf(self, data: bytes) -> str:
        try:
            reader = PdfReader(io.BytesIO(data))
            texts = []
            for page in reader.pages:
                text = page.extract_text() or ""
                if text.strip():
                    texts.append(text.strip())
            return "\n\n".join(texts)
        except Exception as exc:
            logger.exception("Failed to parse PDF")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Failed to parse PDF",
            ) from exc

    def _parse_text(self, data: bytes) -> str:
        try:
            return data.decode("utf-8", errors="ignore")
        except Exception as exc:  # pragma: no cover - defensive
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, detail="Failed to decode text"
            ) from exc

    def _parse_csv(self, data: bytes, delimiter: str = ",") -> str:
        try:
            stream = io.StringIO(data.decode("utf-8", errors="ignore"))
            reader = csv.reader(stream, delimiter=delimiter)
            rows = [" | ".join(row) for row in reader]
            return "\n".join(rows)
        except Exception as exc:
            logger.exception("Failed to parse CSV/TSV")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Failed to parse CSV/TSV",
            ) from exc

    def _parse_excel(self, data: bytes) -> str:
        try:
            wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
            sheets_text = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(cell) for cell in row if cell is not None]
                    if cells:
                        rows.append(" | ".join(cells))
                if rows:
                    sheets_text.append(f"[Sheet: {sheet}]\n" + "\n".join(rows))
            return "\n\n".join(sheets_text)
        except Exception as exc:
            logger.exception("Failed to parse Excel")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Failed to parse Excel file",
            ) from exc

    async def parse_file(self, file: UploadFile, tags: List[str]) -> List[KnowledgeItem]:
        if file.content_type not in self.allowed_types:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Unsupported file type: {file.content_type}",
            )
        data = await file.read()
        if not data:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File is empty")

        text = ""
        ctype = file.content_type
        filename = (file.filename or "").lower()

        if ctype == "application/pdf":
            text = self._parse_pdf(data)
        elif ctype in {"text/plain", "text/markdown"}:
            text = self._parse_text(data)
        elif ctype == "text/csv":
            text = self._parse_csv(data, delimiter=",")
        elif ctype == "text/tab-separated-values":
            text = self._parse_csv(data, delimiter="\t")
        elif ctype == "application/vnd.ms-excel":
            # Disambiguate CSV mislabeled as ms-excel vs real Excel
            if filename.endswith(".csv"):
                text = self._parse_csv(data, delimiter=",")
            else:
                text = self._parse_excel(data)
        elif ctype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            text = self._parse_excel(data)

        text = text.strip()
        if not text:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No text extracted from file")

        chunks = self._chunk_text(text)
        items: List[KnowledgeItem] = []
        for idx, chunk in enumerate(chunks, start=1):
            items.append(
                KnowledgeItem(
                    title=f"{file.filename or 'document'} - part {idx}",
                    content=chunk,
                    tags=tags,
                )
            )
        return items
